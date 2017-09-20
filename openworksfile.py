import sqlite3
from openpyxl import load_workbook

wb = load_workbook(filename = 'worksfile.xlsx')

namasheets=wb.get_sheet_names()

sheet_ranges = wb['wi']
sheet=wb.get_sheet_by_name("wi")
_work=sheet.cell(row=1,column=2).value
_name1=sheet.cell(row=1,column=4).value
_create=sheet.cell(row=2,column=2).value
_name2=sheet.cell(row=2,column=4).value
_lfields=sheet.cell(row=3,column=5).value

rows=sheet.cell(row=4,column=2).value
bts=rows.find(";")

rowsbeginsfrom=rows[:bts]
rowsendsto=rows[bts+1:]
print("rowsbeginsfrom=",rowsbeginsfrom)
print("rowsendsto=",rowsendsto)

columns=sheet.cell(row=5,column=2).value
columnsbeginsfrom=columns[:bts]
columnsendsto=columns[bts+1:]
print("columnsbeginsfrom=",columnsbeginsfrom)
print("columnsendsto=",columnsendsto,"\n")

#### info
##from openpyxl import load_workbook
##
##wb = load_workbook("file.xlsx", data_only=True)
##sh = wb["Sheet_name"]
##print(sh["x10"].value)
##
#### ====
## or cell.internal_value
#### =======================

print("namasheets=",namasheets)
print("Sheet=",sheet)
print("work=",_work)
print("name1=",_name1)
print("create=",_create)
print("name2=",_name2)
print("_lfields=",_lfields)

nbeginfields=7
nlastfield=nbeginfields+_lfields-1
print("nlastfield = ",nlastfield)
i=1

_fields=[sheet.cell(row=3,column=nbeginfields).value]
_tfields=[sheet.cell(row=4,column=nbeginfields).value]
_ftfields=[str(_fields)+" "+str(_tfields)]
print("_fields=",_fields)

while i<=nlastfield-1:
    i+=1
    namafield=sheet.cell(row=3,column=nbeginfields+i).value
    _fields.append(namafield)
    namatype=sheet.cell(row=4,column=nbeginfields+i).value
    _tfields.append(namatype)
    _ftfields.append(str(namafield)+" "+str(namatype))
##    print(i,namafield)

i=0
print("_ftfields=",_ftfields)


while i<=_lfields-1:
    print(i+1,"fields [",i,"]=",_fields[i],",",_tfields[i])
    i+=1
    
i=1
_x="?"
while i<=_lfields-1:
    _x=_x+",?"    
    i+=1
    
print("_x=",_x)

conn = sqlite3.connect(_name1+'.db')
c = conn.cursor()

def readData():
    
    f = open(_name2+'.sql', 'r')
    
    with f:
        data = f.read()
        return data

try:
    data=readData()
    c.commit()
except:
    print("Table not found")

##    Contoh1
##>>> c.execute('''CREATE TABLE example(id INTEGER PRIMARY KEY, created_at DATE)''')
##>>> # Insert a date object into the database
##>>> today = date.today()
##>>> c.execute('''INSERT INTO example(created_at) VALUES(?)''', (today,))

##    Contoh2
##    c.execute('''CREATE TABLE users(id INTEGER PRIMARY KEY, name TEXT, phone TEXT)''')
##    users = [
##    ('John', '5557241'), 
##    ('Adam', '5547874'), 
##    ('Jack', '5484522'), 
##    ('Monthy',' 6656565')
##     ]
## 
##c.executemany('''INSERT INTO users(name, phone) VALUES(?,?)''', users)
##c.executemany('''INSERT INTO table(field, field) VALUES(?,?)''', source)
##perintah="CREATE TABLE _name2("+_x+")"
##    c.execute("CREATE TABLE _name2(?, ?, ?)")

finally:
    
    if conn:
        conn.close() 

    
